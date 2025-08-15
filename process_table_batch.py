import asyncio
import aiohttp
from openpyxl import load_workbook
# from io import BytesIO
# from pydub import AudioSegment
from main import batch_transcribe  # Импортируем функцию для пакетной транскрипции

# Конфигурируемый параметр: размер пакета
BATCH_SIZE = 100
PAUSE_SECONDS = 100

# async def get_audio_info(session: aiohttp.ClientSession, url: str, audio_format="mp3") -> tuple[float, float]:
#     """
#     Загружает аудиофайл по URL и возвращает его длину в секундах и размер в MB.
#     """
#     async with session.get(url) as response:
#         data = await response.read()
#     size_mb = len(data) / (1024 * 1024)
#     audio = AudioSegment.from_file(BytesIO(data), format=audio_format)
#     duration = len(audio) / 1000.0
#     return duration, size_mb

async def process_table_batches(
    excel_file: str, 
    sheet_name: str = None
):
    """
    Обрабатывает Excel‑таблицу группами по 50 записей:
     – выбирает строки, где в столбце E стоит прочерк, 
     – собирает ссылки из столбца F группами по 50 штук,
     – отправляет собранную группу ссылок через groq для транскрипции,
     – записывает транскрипцию в столбец E,
     – в столбец G записывает тег "groq".
    """
    wb = load_workbook(excel_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Сбор строк для обработки
    rows_to_process = []
    for row in ws.iter_rows(min_row=2):
        cell_E = row[4]  # столбец E
        cell_F = row[5]  # столбец F
        if cell_E.value == "-" and cell_F.value:
            rows_to_process.append(row)

    print(f"Найдено {len(rows_to_process)} записей для обработки")

    # Разбиваем на группы по BATCH_SIZE записей
    total_processed = 0
    
    for i in range(0, len(rows_to_process), BATCH_SIZE):
        batch_rows = rows_to_process[i:i + BATCH_SIZE]
        batch_urls = [row[5].value for row in batch_rows]
        
        print(f"Обработка группы {i//BATCH_SIZE + 1}, записей: {len(batch_urls)}")
        
        results = await batch_transcribe(batch_urls)
        
        # Обновляем таблицу
        for res, proc_row in zip(results, batch_rows):
            if res["status"] == "success":
                proc_row[4].value = res["text"]   # Обновляем столбец E транскрипцией
                proc_row[6].value = "groq"        # В столбец G записываем тег "groq"
                total_processed += 1
            else:
                proc_row[4].value = f"Error: {res['error']}"
        
        # Сохраняем промежуточные результаты
        wb.save(excel_file)
        print(f"Промежуточное сохранение файла '{excel_file}'")
        
        # Пауза перед следующей группой
        if i + BATCH_SIZE < len(rows_to_process):
            print(f"Пауза {PAUSE_SECONDS} секунд перед следующей группой...")
            await asyncio.sleep(PAUSE_SECONDS)

    print(f"Файл '{excel_file}' обновлен.")
    print(f"Всего успешно обработано {total_processed} записей из {len(rows_to_process)}")

if __name__ == "__main__":
    import sys
    excel_file = "DFASDF.xlsx"  # Замените на актуальный путь
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    asyncio.run(process_table_batches(excel_file)) 