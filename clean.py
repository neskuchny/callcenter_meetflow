import openpyxl

def clean_excel_file(filename):
    # Загружаем рабочую книгу
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    print("Начинаем обработку файла...")
    # Проходим по всем строкам в листе
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        # Проверяем значения в столбцах E и G (индексы 4 и 6)
        e_value = str(row[4].value).strip() if row[4].value is not None else ''
        g_value = str(row[6].value).strip() if row[6].value is not None else ''
        
        print(f"Строка {row_idx}: E='{e_value}', G='{g_value}'")
        
        if g_value == '':
            # Если ячейки в столбцах E и G пусты (после удаления пробелов), ставим "-" в столбце E
            row[4].value = "-"
            print(f"ИЗМЕНЕНО: Строка {row_idx} - установлен символ '-' в столбце E")

    print("Сохраняем изменения...")
    output_filename = filename.replace('.xlsx', '_clean.xlsx')
    wb.save(output_filename)
    print("Готово!")

def remove_dash_rows(filename):
    print("Начинаем удаление строк со значением '-' в столбце E...")
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    # Идем с конца, чтобы удаление строк не влияло на индексы
    rows_to_delete = []
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        e_value = str(row[4].value).strip() if row[4].value is not None else ''
        if e_value == '-':
            rows_to_delete.append(row_idx)
            print(f"Найдена строка {row_idx} для удаления")
    
    # Удаляем строки в обратном порядке
    for row_idx in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_idx)
        print(f"Удалена строка {row_idx}")
    
    print(f"Всего удалено строк: {len(rows_to_delete)}")
    output_filename = filename.replace('.xlsx', '_without_dash.xlsx')
    wb.save(output_filename)
    print("Готово!")

if __name__ == "__main__":
    # clean_excel_file('DFASDF.xlsx')
    remove_dash_rows('DFASDF_clean.xlsx')
